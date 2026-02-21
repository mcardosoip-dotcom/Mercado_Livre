"""
Script para Gerar Arquivo Excel com Filtros a partir da tabela LK_PBD_LA_ENTRADAS_E_DESFECHOS
Autor: Marcelo Cardoso
Data: 2026-02-20

Objetivo:
    A consulta é feita na tabela BigQuery; os filtros ficam no WHERE da query.
    O BigQuery retorna apenas o resultado já filtrado; o script só grava esse resultado no Excel.
    Nenhuma filtragem é feita em Python.

Entrada: tabela BigQuery (pdme000426-c1s7scatwm0-furyid.STG.LK_PBD_LA_ENTRADAS_E_DESFECHOS)
Saída: arquivo Excel (output_bigquery.xlsx)

Filtros (aplicados na query, na tabela):
    - area_do_direito = "Requerimentos"
    - sub_area_do_direito = "Ofícios"
    - DATA_REGISTRADO_TRATADA = Agosto/2025
    - processo_materia = "EXJUD"
"""

import os
import warnings
import pandas as pd
from datetime import datetime
from openpyxl.utils import get_column_letter

# Suprime aviso de resultado grande quando filtros já estão na query
try:
    import pandas_gbq.exceptions as _gbq_exc
    warnings.simplefilter("ignore", category=_gbq_exc.LargeResultsWarning)
except Exception:
    pass

# ================================================
# CONFIGURAÇÕES
# ================================================

# Pasta raiz do script
PASTA_RAIZ_SCRIPT = os.path.dirname(os.path.abspath(__file__))

# Entrada: tabela BigQuery
BQ_PROJETO = "pdme000426-c1s7scatwm0-furyid"
BQ_DATASET = "STG"
BQ_TABELA = "LK_PBD_LA_ENTRADAS_E_DESFECHOS"

# Saída: arquivo Excel
NOME_ARQUIVO_EXCEL = "output_bigquery.xlsx"

# ================================================
# CONSULTA NA TABELA (retorna apenas o resultado filtrado)
# ================================================

def _query_bigquery_filtrada():
    """Query na tabela com filtros no WHERE. O BigQuery retorna apenas as linhas que atendem aos critérios."""
    table_id = f"{BQ_PROJETO}.{BQ_DATASET}.{BQ_TABELA}"
    where = """
    WHERE LOWER(TRIM(SAFE_CAST(area_do_direito AS STRING))) = 'requerimentos'
      AND LOWER(TRIM(SAFE_CAST(sub_area_do_direito AS STRING))) = 'ofícios'
      AND DATA_REGISTRADO_TRATADA >= DATE '2025-08-01'
      AND DATA_REGISTRADO_TRATADA < DATE '2025-09-01'
      AND TRIM(SAFE_CAST(processo_materia AS STRING)) = 'EXJUD'
    """
    return f"SELECT * FROM `{table_id}` {where}"


def carregar_dados_bigquery():
    """Executa a consulta na tabela BigQuery; retorna apenas o resultado já filtrado (WHERE na query)."""
    try:
        import pandas_gbq
    except ImportError:
        print("   [ERRO] Modo BigQuery requer: pip install pandas-gbq pyarrow db-dtypes")
        return None
    table_id = f"{BQ_PROJETO}.{BQ_DATASET}.{BQ_TABELA}"
    print(f"   [INFO] Consulta na tabela (filtros no WHERE): {table_id}")
    query = _query_bigquery_filtrada()
    return pandas_gbq.read_gbq(
        query,
        project_id=BQ_PROJETO,
        progress_bar_type=None,
    )


def criar_arquivo_excel_com_filtros(df_final, caminho_excel):
    """Grava no Excel o resultado retornado pela consulta (sem filtragem em Python)."""
    if df_final is None or df_final.empty:
        return False, 0, 0

    df_exportar = df_final
    print(f"\n[CRIANDO] Arquivo Excel: {NOME_ARQUIVO_EXCEL}")

    with pd.ExcelWriter(caminho_excel, engine="openpyxl") as writer:
        nome_aba = "LK_PBD_Entradas_Desfechos"[:31]
        df_exportar.to_excel(writer, sheet_name=nome_aba, index=False)
        worksheet = writer.sheets[nome_aba]

        num_linhas = len(df_exportar) + 1
        num_colunas = len(df_exportar.columns)
        coluna_final_letra = get_column_letter(num_colunas)
        worksheet.auto_filter.ref = f"A1:{coluna_final_letra}{num_linhas}"

        for idx, col in enumerate(df_exportar.columns, 1):
            max_length = max(
                len(str(col)),
                df_exportar[col].astype(str).map(len).max() if len(df_exportar) > 0 else 0,
            )
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[worksheet.cell(row=1, column=idx).column_letter].width = adjusted_width

    print(f"   [OK] Arquivo salvo: {caminho_excel}")
    print(f"   [INFO] Linhas exportadas: {len(df_exportar):,} | Colunas: {len(df_exportar.columns)}")
    return True, len(df_exportar), len(df_exportar.columns)


# ================================================
# MAIN
# ================================================

def main():
    print("=" * 70)
    print("EXCEL COM FILTROS - FONTE: LK_PBD_LA_ENTRADAS_E_DESFECHOS")
    print("=" * 70)
    print(f"Data/Hora: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    print()

    caminho_excel = os.path.join(PASTA_RAIZ_SCRIPT, NOME_ARQUIVO_EXCEL)

    print("[LENDO] BigQuery - LK_PBD_LA_ENTRADAS_E_DESFECHOS")
    df = carregar_dados_bigquery()

    if df is None:
        print("\n[ERRO] Nao foi possivel carregar os dados da tabela.")
        return

    print(f"   [OK] Linhas: {len(df):,} | Colunas: {len(df.columns)}")

    sucesso, num_linhas, num_colunas = criar_arquivo_excel_com_filtros(df, caminho_excel)

    print("\n" + "=" * 70)
    print("RESUMO FINAL")
    print("=" * 70)
    if sucesso:
        print(f"[OK] Arquivo criado: {caminho_excel}")
        print(f"[TOTAL] Linhas: {num_linhas:,} | Colunas: {num_colunas}")
        print("\nCompare com output_parquet.xlsx para verificar se o resultado e o mesmo.")
    else:
        print("[ERRO] Falha ao criar Excel.")
    print("\n[SUCESSO] Processo concluido.")


if __name__ == "__main__":
    main()
