"""
Script para Gerar Arquivo Excel com Filtros do Consolidado
Autor: Marcelo Cardoso
Data: 2026-01-06

Objetivo:
    Ler o arquivo parquet consolidado, aplicar filtros específicos e gerar um arquivo Excel
    com filtros automáticos habilitados em todas as colunas.
    
Arquivo base:
    Database_eLAW_Contencioso_Consolidado_Final.parquet
    (localizado na mesma pasta do script: Embargos)
    
Filtros aplicados:
    - area_do_direito = "Requerimentos"
    - sub_area_do_direito = "Ofícios"
    - data_registrado = "Agosto de 2025"
    - processo_materia = "EXJUD"
"""

import os
import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl.utils import get_column_letter

# ================================================
# CONFIGURAÇÕES
# ================================================

# Pasta raiz do script: Onde o arquivo Python está localizado (pasta Embargos)
PASTA_RAIZ_SCRIPT = os.path.dirname(os.path.abspath(__file__))

# Nome do arquivo parquet consolidado a ser lido
# Arquivo base: Database_eLAW_Contencioso_Consolidado_Final.parquet
# Localizado em: H:\Drives compartilhados\Legales_Analytics\002 - Python\002-1 - DEV\Dev Marcelo\0000\Confronto\Embargos\
NOME_ARQUIVO_PARQUET = "Database_eLAW_Contencioso_Consolidado_Final.parquet"

# Nome do arquivo Excel de saída
NOME_ARQUIVO_EXCEL = "output_parquet.xlsx"

# ================================================
# FUNÇÕES AUXILIARES
# ================================================

def encontrar_coluna(df, nome_coluna):
    """
    Encontra uma coluna no DataFrame de forma case-insensitive.
    
    Args:
        df: DataFrame para buscar a coluna
        nome_coluna: Nome da coluna a buscar (case-insensitive)
    
    Returns:
        Nome da coluna encontrada ou None
    """
    nome_coluna_lower = nome_coluna.lower()
    for col in df.columns:
        if col.lower() == nome_coluna_lower:
            return col
    return None


def aplicar_filtros(df):
    """
    Aplica os filtros específicos no DataFrame:
    - area_do_direito = "Requerimentos"
    - sub_area_do_direito = "Ofícios"
    - data_registrado = "Agosto de 2025"
    - processo_materia = "EXJUD"
    
    Args:
        df: DataFrame para filtrar
    
    Returns:
        DataFrame filtrado
    """
    df_filtrado = df.copy()
    linhas_antes = len(df_filtrado)
    
    print(f"\n[APLICANDO] Filtros nos dados...")
    print(f"   [INFO] Linhas antes dos filtros: {linhas_antes:,}")
    
    # Encontra as colunas necessárias
    col_area_do_direito = encontrar_coluna(df_filtrado, "area_do_direito")
    col_sub_area_do_direito = encontrar_coluna(df_filtrado, "sub_area_do_direito")
    col_data_registrado = encontrar_coluna(df_filtrado, "data_registrado")
    col_processo_materia = encontrar_coluna(df_filtrado, "processo_materia")
    
    # Debug: mostra informações sobre as colunas encontradas
    print(f"\n   [DEBUG] Colunas encontradas:")
    print(f"      - area_do_direito: {col_area_do_direito if col_area_do_direito else 'NAO ENCONTRADA'}")
    print(f"      - sub_area_do_direito: {col_sub_area_do_direito if col_sub_area_do_direito else 'NAO ENCONTRADA'}")
    print(f"      - data_registrado: {col_data_registrado if col_data_registrado else 'NAO ENCONTRADA'}")
    print(f"      - processo_materia: {col_processo_materia if col_processo_materia else 'NAO ENCONTRADA'}")
    
    # Aplica filtro de area_do_direito
    if col_area_do_direito:
        # Mostra valores únicos antes de filtrar
        valores_unicos = df_filtrado[col_area_do_direito].dropna().unique()[:10]
        print(f"\n   [FILTRO] area_do_direito = 'Requerimentos'")
        print(f"      [DEBUG] Valores unicos encontrados (primeiros 10): {list(valores_unicos)}")
        
        # Filtra de forma case-insensitive e remove espaços
        mask = df_filtrado[col_area_do_direito].astype(str).str.strip().str.lower() == "requerimentos"
        df_filtrado = df_filtrado[mask]
        print(f"      [OK] Linhas apos filtro area_do_direito: {len(df_filtrado):,}")
    else:
        print(f"   [ERRO] Coluna 'area_do_direito' nao encontrada. Pulando filtro.")
        print(f"   [INFO] Colunas disponiveis (primeiras 30): {list(df_filtrado.columns)[:30]}")
        # Tenta encontrar colunas similares
        colunas_similares = [col for col in df_filtrado.columns if "area" in col.lower() or "direito" in col.lower()]
        if colunas_similares:
            print(f"   [INFO] Colunas similares encontradas: {colunas_similares}")
    
    # Aplica filtro de sub_area_do_direito
    if col_sub_area_do_direito:
        # Mostra valores únicos antes de filtrar
        valores_unicos = df_filtrado[col_sub_area_do_direito].dropna().unique()[:10]
        print(f"\n   [FILTRO] sub_area_do_direito = 'Ofícios'")
        print(f"      [DEBUG] Valores unicos encontrados (primeiros 10): {list(valores_unicos)}")
        
        # Filtra de forma case-insensitive e remove espaços
        mask = df_filtrado[col_sub_area_do_direito].astype(str).str.strip().str.lower() == "ofícios"
        df_filtrado = df_filtrado[mask]
        print(f"      [OK] Linhas apos filtro sub_area_do_direito: {len(df_filtrado):,}")
    else:
        print(f"   [AVISO] Coluna 'sub_area_do_direito' nao encontrada. Pulando filtro.")
    
    # Aplica filtro de data_registrado = "Agosto de 2025"
    if col_data_registrado:
        # Mostra valores únicos antes de filtrar
        valores_unicos = df_filtrado[col_data_registrado].dropna().unique()[:10]
        tipo_dados = df_filtrado[col_data_registrado].dtype
        print(f"\n   [FILTRO] data_registrado = 'Agosto de 2025'")
        print(f"      [DEBUG] Tipo de dados da coluna: {tipo_dados}")
        print(f"      [DEBUG] Valores unicos encontrados (primeiros 10): {list(valores_unicos)}")
        print(f"      [DEBUG] Exemplo de valores (primeiros 5): {list(df_filtrado[col_data_registrado].dropna().head(5).values)}")
        
        # Tenta múltiplas estratégias de filtro
        mask = None
        
        # Estratégia 1: Se for datetime ou puder ser convertido para datetime
        try:
            # Tenta converter para datetime
            df_filtrado_temp = df_filtrado.copy()
            df_filtrado_temp[col_data_registrado + '_dt'] = pd.to_datetime(
                df_filtrado_temp[col_data_registrado], 
                errors='coerce',
                dayfirst=True,
                format='mixed'
            )
            
            # Filtra por ano 2025 e mês agosto (8)
            mask_ano = df_filtrado_temp[col_data_registrado + '_dt'].dt.year == 2025
            mask_mes = df_filtrado_temp[col_data_registrado + '_dt'].dt.month == 8
            
            if mask_ano.any() and mask_mes.any():
                mask = mask_ano & mask_mes
                print(f"      [INFO] Filtro aplicado usando conversao para datetime (ano=2025, mes=8)")
        except Exception as e:
            print(f"      [DEBUG] Conversao para datetime falhou: {e}")
        
        # Estratégia 2: Se a conversão para datetime não funcionou ou não encontrou resultados, tenta filtro por string
        if mask is None or (hasattr(mask, 'any') and not mask.any()):
            print(f"      [INFO] Tentando filtro por string...")
            col_str = df_filtrado[col_data_registrado].astype(str).str.strip().str.lower()
            
            # Procura por vários padrões possíveis
            mask = (
                # "Agosto de 2025" (formato completo)
                col_str.str.contains("agosto.*2025", case=False, na=False, regex=True) |
                # "2025-08" (formato ISO)
                col_str.str.contains("2025-08", case=False, na=False) |
                # "08/2025" ou "08-2025" (formato mês/ano)
                col_str.str.contains("08[/-]2025", case=False, na=False, regex=True) |
                # "08/2025" em qualquer ordem (mas ambos devem estar presentes)
                (col_str.str.contains("2025", case=False, na=False) & col_str.str.contains("08", case=False, na=False)) |
                # "august 2025" (inglês)
                col_str.str.contains("august.*2025", case=False, na=False, regex=True)
            )
            print(f"      [INFO] Filtro aplicado usando busca por string")
            print(f"      [DEBUG] Linhas que correspondem ao padrao: {mask.sum():,} de {len(mask):,}")
        
        # Aplica o filtro
        if mask is not None and hasattr(mask, 'any'):
            linhas_antes_data = len(df_filtrado)
            df_filtrado = df_filtrado[mask]
            linhas_depois_data = len(df_filtrado)
            print(f"      [OK] Linhas apos filtro data_registrado: {linhas_depois_data:,} (removidas: {linhas_antes_data - linhas_depois_data:,})")
            
            # Mostra alguns exemplos dos valores que passaram no filtro
            if len(df_filtrado) > 0:
                exemplos = df_filtrado[col_data_registrado].dropna().unique()[:5]
                print(f"      [DEBUG] Exemplos de valores que passaram no filtro: {list(exemplos)}")
            else:
                print(f"      [AVISO] Nenhuma linha passou no filtro de data!")
                print(f"      [DEBUG] Verifique se os valores na coluna correspondem a 'Agosto de 2025' ou formato equivalente")
        else:
            print(f"      [ERRO] Nao foi possivel criar mascara de filtro para data_registrado")
    else:
        print(f"   [AVISO] Coluna 'data_registrado' nao encontrada. Pulando filtro.")
    
    # Aplica filtro de processo_materia
    if col_processo_materia:
        # Mostra valores únicos antes de filtrar
        valores_unicos = df_filtrado[col_processo_materia].dropna().unique()[:10]
        print(f"\n   [FILTRO] processo_materia = 'EXJUD'")
        print(f"      [DEBUG] Valores unicos encontrados (primeiros 10): {list(valores_unicos)}")
        
        # Filtra por string exata (case-sensitive para códigos)
        mask = df_filtrado[col_processo_materia].astype(str).str.strip() == "EXJUD"
        df_filtrado = df_filtrado[mask]
        print(f"      [OK] Linhas apos filtro processo_materia: {len(df_filtrado):,}")
    else:
        print(f"   [AVISO] Coluna 'processo_materia' nao encontrada. Pulando filtro.")
    
    linhas_depois = len(df_filtrado)
    linhas_removidas = linhas_antes - linhas_depois
    
    print(f"\n   [RESUMO] Linhas antes: {linhas_antes:,}")
    print(f"   [RESUMO] Linhas depois: {linhas_depois:,}")
    print(f"   [RESUMO] Linhas removidas: {linhas_removidas:,}")
    
    # Verifica se alguma coluna não foi encontrada
    colunas_nao_encontradas = []
    if not col_area_do_direito:
        colunas_nao_encontradas.append("area_do_direito")
    if not col_sub_area_do_direito:
        colunas_nao_encontradas.append("sub_area_do_direito")
    if not col_data_registrado:
        colunas_nao_encontradas.append("data_registrado")
    if not col_processo_materia:
        colunas_nao_encontradas.append("processo_materia")
    
    if colunas_nao_encontradas:
        print(f"\n   [ATENCAO] As seguintes colunas nao foram encontradas: {colunas_nao_encontradas}")
        print(f"   [INFO] Todas as colunas disponiveis no arquivo:")
        # Usa o DataFrame original antes dos filtros
        df_original = df.copy()
        for i, col in enumerate(df_original.columns, 1):
            print(f"      {i}. {col}")
    
    return df_filtrado


# ================================================
# FUNÇÕES PRINCIPAIS
# ================================================

def criar_arquivo_excel_com_filtros(caminho_parquet, caminho_excel):
    """
    Cria um arquivo Excel com os dados consolidados e filtros automáticos habilitados.
    
    Args:
        caminho_parquet: Caminho do arquivo parquet consolidado
        caminho_excel: Caminho onde salvar o arquivo Excel
    """
    try:
        # Verifica se o arquivo parquet existe
        if not os.path.exists(caminho_parquet):
            print(f"   [ERRO] Arquivo parquet nao encontrado: {caminho_parquet}")
            return False, 0, 0
        
        print(f"\n[LENDO] Arquivo parquet: {os.path.basename(caminho_parquet)}")
        df_final = pd.read_parquet(caminho_parquet)
        print(f"   [OK] Total de linhas: {len(df_final):,}")
        print(f"   [OK] Total de colunas: {len(df_final.columns)}")
        
        if df_final.empty:
            print(f"   [AVISO] DataFrame vazio, nao e possivel criar arquivo Excel com filtros.")
            return False, 0, 0
        
        # Aplica os filtros específicos
        df_exportar = aplicar_filtros(df_final)
        
        if df_exportar.empty:
            print(f"   [AVISO] Nenhum registro encontrado apos aplicar os filtros.")
            print(f"   [INFO] Verifique os criterios de filtro:")
            print(f"      - area_do_direito = 'Requerimentos'")
            print(f"      - sub_area_do_direito = 'Ofícios'")
            print(f"      - data_registrado = 'Agosto de 2025'")
            print(f"      - processo_materia = 'EXJUD'")
            return False, 0, 0
        
        print(f"\n[CRIANDO] Arquivo Excel com filtros: {NOME_ARQUIVO_EXCEL}")
        
        # Salva em Excel
        with pd.ExcelWriter(caminho_excel, engine='openpyxl') as writer:
            # Cria nome da aba (limita a 31 caracteres)
            nome_aba = "Dados_Consolidados"
            if len(nome_aba) > 31:
                nome_aba = nome_aba[:31]
            
            # Exporta o DataFrame
            df_exportar.to_excel(writer, sheet_name=nome_aba, index=False)
            
            # Obtém a planilha para aplicar filtros
            worksheet = writer.sheets[nome_aba]
            
            # Aplica filtros automáticos na primeira linha (cabeçalho)
            if len(df_exportar) > 0:
                # Define a faixa de dados (incluindo cabeçalho)
                # Linha final = número de linhas de dados + 1 (cabeçalho)
                num_linhas = len(df_exportar) + 1
                num_colunas = len(df_exportar.columns)
                
                # Converte número da coluna em letra (ex: 1 -> A, 26 -> Z, 27 -> AA)
                coluna_final_letra = get_column_letter(num_colunas)
                
                start_cell_address = 'A1'
                end_cell_address = f"{coluna_final_letra}{num_linhas}"
                
                # Aplica o filtro automático
                worksheet.auto_filter.ref = f"{start_cell_address}:{end_cell_address}"
                
                # Ajusta a largura das colunas automaticamente
                for idx, col in enumerate(df_exportar.columns, 1):
                    # Calcula a largura máxima do conteúdo
                    max_length = max(
                        len(str(col)),  # Largura do cabeçalho
                        df_exportar[col].astype(str).map(len).max() if len(df_exportar) > 0 else 0  # Largura do conteúdo
                    )
                    # Limita a largura máxima a 50 caracteres para não ficar muito largo
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[worksheet.cell(row=1, column=idx).column_letter].width = adjusted_width
                
                print(f"   [OK] Filtros automáticos aplicados na faixa: {start_cell_address}:{end_cell_address}")
            
            print(f"   [OK] Aba '{nome_aba}' criada com sucesso!")
        
        print(f"   [OK] Arquivo Excel com filtros criado com sucesso: {caminho_excel}")
        num_linhas = len(df_exportar)
        num_colunas = len(df_exportar.columns)
        print(f"   [INFO] Total de linhas exportadas: {num_linhas:,}")
        print(f"   [INFO] Total de colunas exportadas: {num_colunas}")
        return True, num_linhas, num_colunas
        
    except FileNotFoundError:
        print(f"   [ERRO] Arquivo parquet nao encontrado: {caminho_parquet}")
        return False, 0, 0
    except Exception as e:
        print(f"   [ERRO] Erro ao criar arquivo Excel com filtros: {e}")
        import traceback
        traceback.print_exc()
        return False, 0, 0


# ================================================
# FUNÇÃO PRINCIPAL
# ================================================

def main():
    """Função principal que executa o processo de geração do Excel com filtros."""
    
    print("=" * 70)
    print("GERAÇÃO DE ARQUIVO EXCEL COM FILTROS")
    print("=" * 70)
    print(f"Data/Hora: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    print()
    
    # Define os caminhos
    caminho_parquet = os.path.join(PASTA_RAIZ_SCRIPT, NOME_ARQUIVO_PARQUET)
    caminho_excel = os.path.join(PASTA_RAIZ_SCRIPT, NOME_ARQUIVO_EXCEL)
    
    print(f"[PASTA] Script: {PASTA_RAIZ_SCRIPT}")
    print(f"[ARQUIVO] Parquet de entrada: {NOME_ARQUIVO_PARQUET}")
    print(f"[ARQUIVO] Excel de saida: {NOME_ARQUIVO_EXCEL}")
    
    # Executa a criação do arquivo Excel
    sucesso, num_linhas_exportadas, num_colunas_exportadas = criar_arquivo_excel_com_filtros(caminho_parquet, caminho_excel)
    
    # ============================================
    # RESUMO FINAL
    # ============================================
    print("\n" + "=" * 70)
    print("RESUMO FINAL")
    print("=" * 70)
    
    if sucesso:
        print(f"[OK] Arquivo Excel criado com sucesso!")
        print(f"[ARQUIVO] {caminho_excel}")
        print(f"[TOTAL] Linhas exportadas no arquivo Excel: {num_linhas_exportadas:,}")
        print(f"[TOTAL] Colunas exportadas no arquivo Excel: {num_colunas_exportadas}")
    else:
        print(f"[ERRO] Falha ao criar arquivo Excel.")
        print(f"[INFO] Verifique se o arquivo parquet existe: {caminho_parquet}")
        if num_linhas_exportadas == 0:
            print(f"[INFO] Nenhuma linha foi exportada (arquivo vazio ou filtros muito restritivos)")
    
    print("\n[SUCESSO] Processo concluido!")


if __name__ == "__main__":
    main()
