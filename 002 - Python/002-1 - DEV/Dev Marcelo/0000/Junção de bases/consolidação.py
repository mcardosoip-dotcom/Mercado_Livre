
import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

PASTA_ORIGEM = Path(r"C:\Users\mcard\Desktop\Temp\Legado")
NOME_ABA = "Extração Marcelo (contencioso2)"
LINHAS_CABECALHO = 5  # linhas a manter no topo, desconsideradas na junção
DATA_LIMITE = datetime(2026, 1, 1)  # manter apenas registros com Data Registrado antes desta data
COLUNA_DATA_REGISTRADO = "Data Registrado"  # ou coluna AG (índice 32) se o nome não existir

# Teste: limitar resultado a N linhas e salvar em arquivo _TESTE (altere para None para rodar completo)
TESTE_LIMITE_LINHAS = None  # 100 = teste com 100 linhas; None = base full


def ler_cabecalho_5_linhas(caminho: Path):
    """Lê as primeiras 5 linhas do primeiro sheet para manter no resultado."""
    wb = load_workbook(caminho, read_only=True, data_only=True)
    ws = wb.active
    linhas = []
    for row in ws.iter_rows(min_row=1, max_row=LINHAS_CABECALHO, values_only=True):
        linhas.append(list(row))
    wb.close()
    return linhas


def ler_dados_excel(caminho: Path, nrows: int | None = None) -> pd.DataFrame:
    """Lê dados do Excel: linhas 1-5 = cabeçalho decorativo, linha 6 = nomes das colunas, 7+ = dados."""
    kwargs = {"header": LINHAS_CABECALHO, "engine": "openpyxl"}  # linha 6 = índice 5 (0-based)
    if nrows is not None:
        kwargs["nrows"] = nrows
    return pd.read_excel(caminho, **kwargs)


def consolidar_grupo(
    arquivo1: Path,
    arquivo2: Path,
    arquivo_saida: Path,
    coluna_dedup: str = "(Processo) ID",
) -> Path:
    """
    Junta dois arquivos do mesmo grupo, mantém cabeçalho de 5 linhas,
    remove duplicatas por coluna_dedup e grava um único xlsx.
    """
    # Cabecalho das 5 linhas (do primeiro arquivo)
    cabecalho_linhas = ler_cabecalho_5_linhas(arquivo1)

    # Dados dos dois arquivos (a partir da linha 6); em modo teste lê menos linhas
    nrows = TESTE_LIMITE_LINHAS if TESTE_LIMITE_LINHAS else None
    nrows_por_arquivo = ((nrows // 2) + 10) if nrows else None  # lê um pouco a mais para ter ~100 após filtros
    df1 = ler_dados_excel(arquivo1, nrows=nrows_por_arquivo)
    df2 = ler_dados_excel(arquivo2, nrows=nrows_por_arquivo)
    df = pd.concat([df1, df2], ignore_index=True)

    # Garantir que os tipos de dado sejam os mesmos do primeiro arquivo
    for col in df1.columns:
        if col in df.columns and df1[col].dtype != df[col].dtype:
            try:
                df[col] = df[col].astype(df1[col].dtype, errors="ignore")
            except (TypeError, ValueError):
                pass

    # Filtro: apenas Data Registrado antes de 01/01/2026 (coluna "Data Registrado" ou AG)
    if COLUNA_DATA_REGISTRADO in df.columns:
        col_data = COLUNA_DATA_REGISTRADO
    elif len(df.columns) > 32:
        col_data = df.columns[32]  # coluna AG (índice 32)
    else:
        col_data = None
    if col_data is not None:
        df[col_data] = pd.to_datetime(df[col_data], errors="coerce")
        df = df[df[col_data] < DATA_LIMITE]

    # Remover duplicatas por (Processo) ID (mantém primeira ocorrência)
    if coluna_dedup in df.columns:
        df = df.drop_duplicates(subset=[coluna_dedup], keep="first")
    else:
        # fallback: tentar por "Processo" ou "ID"
        for col in ["Processo", "ID"]:
            if col in df.columns:
                df = df.drop_duplicates(subset=[col], keep="first")
                break

    # Modo teste: limitar a N linhas no resultado
    if TESTE_LIMITE_LINHAS:
        df = df.head(TESTE_LIMITE_LINHAS)

    # Gravar: primeiro as 5 linhas de cabeçalho, depois o dataframe
    with pd.ExcelWriter(arquivo_saida, engine="openpyxl") as writer:
        df.to_excel(
            writer,
            sheet_name=NOME_ABA,
            startrow=LINHAS_CABECALHO,
            index=False,
        )

    # Inserir as 5 linhas de cabeçalho no topo do arquivo gerado
    wb = load_workbook(arquivo_saida)
    ws = wb[NOME_ABA]
    for i, row in enumerate(cabecalho_linhas, start=1):
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=value)
    wb.save(arquivo_saida)
    wb.close()
    return arquivo_saida


def main():
    origem = PASTA_ORIGEM
    sufixo = "_TESTE_100" if TESTE_LIMITE_LINHAS else ""

    # Grupo Incoming
    consolidar_grupo(
        arquivo1=origem / "Database_eLAW_Contencioso_Hispanos_Incoming.xlsx",
        arquivo2=origem / "Database_eLAW_Contencioso_Hispanos_Incoming_2.xlsx",
        arquivo_saida=origem / f"Consolidado_eLAW_Contencioso_Hispanos_Incoming{sufixo}.xlsx",
    )
    print("Gerado:", origem / f"Consolidado_eLAW_Contencioso_Hispanos_Incoming{sufixo}.xlsx")

    # Grupo Outgoing
    consolidar_grupo(
        arquivo1=origem / "Database_eLAW_Contencioso_Hispanos_Outgoing.xlsx",
        arquivo2=origem / "Database_eLAW_Contencioso_Hispanos_Outgoing_2.xlsx",
        arquivo_saida=origem / f"Consolidado_eLAW_Contencioso_Hispanos_Outgoing{sufixo}.xlsx",
    )
    print("Gerado:", origem / f"Consolidado_eLAW_Contencioso_Hispanos_Outgoing{sufixo}.xlsx")

    if TESTE_LIMITE_LINHAS:
        print(f"Concluído: 2 arquivos de TESTE (máx. {TESTE_LIMITE_LINHAS} linhas cada).")
    else:
        print("Concluído: 2 arquivos gerados.")


if __name__ == "__main__":
    main()
